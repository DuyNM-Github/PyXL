import os
import re
from copy import deepcopy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException


class ExcelHandler:
    """
        Reusable chunk of codes that focus on the handling of Excel files
    """

    wb = Workbook()

    def __init__(self):
        self.active_workbook = None
        self.active_workbook_filepath = None
        self.active_workbook_data = None
        self.loaded_workbooks = {}
        self.alias_dict = {}
        self.save_storage = {}

    def open_workbook(self, path_to_file, alias=None):
        try:
            temp1 = load_workbook(path_to_file)
            temp2 = load_workbook(path_to_file, data_only=True, read_only=True)
            self.active_workbook = temp1
            if alias is not None:
                self.loaded_workbooks[alias] = temp1
                self.loaded_workbooks[f"{alias}_data"] = temp2
                self.alias_dict[alias] = path_to_file
            else:
                self.loaded_workbooks[os.path.basename(path_to_file)] = temp1
                self.loaded_workbooks[f"{os.path.basename(path_to_file)}_data"] = temp2
                self.alias_dict[os.path.basename(path_to_file)] = path_to_file
            self.active_workbook_filepath = path_to_file
            print(f"Excel file {os.path.basename(path_to_file)} loaded")
        except (InvalidFileException, FileNotFoundError) as error:
            print(f"FILE LOADING ERROR: {error}")

    def close_workbook(self, file_name=None, alias=None):
        if alias is not None:
            if self.get_active_workbook() == self.loaded_workbooks[alias]:
                self.active_workbook = None
            del (self.loaded_workbooks[alias])
            del (self.alias_dict[alias])
            print(f"Workbook with alias \'{alias}\' closed")
        elif file_name is not None:
            if "/" in file_name or "\\" in file_name:
                file_name = os.path.basename(file_name)
            if self.get_active_workbook() == self.loaded_workbooks[file_name]:
                self.active_workbook = None
            del (self.loaded_workbooks[os.path.basename(file_name)])
            del (self.alias_dict[file_name])
            print(f"Workbook with alias \'{os.path.basename(file_name)}\' closed")
        elif file_name is None and alias is None:
            print("Specify the path to file or the alias associated to the workbook")

    def switch_workbook(self, file_name=None, alias=None, suppress_msg=False):
        for key, values in self.loaded_workbooks.items():
            if file_name is None and alias is None:
                print("Passed in file name and alias is None.")
                return
            if key == file_name or key == alias:
                self.active_workbook = values
                self.active_workbook_data = self.loaded_workbooks[f"{key}_data"]
                self.active_workbook_filepath = self.alias_dict[key]
                if suppress_msg is not False:
                    print(f"Active workbook switched to {key}")
                return
        print("WORKBOOK SWITCHING ERROR: Requested workbook is not found")
        print(f"Hint. Loaded workbooks: {self.loaded_workbooks.keys()}")

    def set_active_sheet(self, sheet_name=None):
        wb = self.active_workbook
        if wb is None:
            print("EXCEL ERROR: No Excel file is currently loaded")
            return None
        if sheet_name is not None:
            for s in range(len(wb.sheetnames)):
                if wb.sheetnames[s] == sheet_name:
                    wb.active = s
                    active_sheet = wb.active
                    print(f"Active sheet switched to {sheet_name}")
                    return active_sheet
            print(f"EXCEL ERROR: Sheet {sheet_name} not found")
            return None

    def get_active_sheet(self):
        return self.active_workbook.active

    def get_loaded_workbooks(self):
        return self.loaded_workbooks

    def get_active_workbook(self):
        return self.active_workbook

    def save_active_workbook(self):
        wb = self.active_workbook
        try:
            wb.save(self.active_workbook_filepath)
        except PermissionError:
            print(f"Error SAVING FILE: File {self.active_workbook_filepath} is opened in another process")

    def save_all_workbook(self):
        current_active_excel_file = self.active_workbook
        for key, value in self.alias_dict.items():
            self.switch_workbook(key, suppress_msg=True)
            try:
                self.save_active_workbook()
            except PermissionError:
                print(f"ERROR SAVING FILE: File {value} is opened in another process")
                pass
        self.active_workbook = current_active_excel_file

    def __read_sheet(self):
        pass

    def get_sheet_value(self, sheet_range, data_only=False):
        if data_only is True:
            ws = self.active_workbook_data.active
        else:
            ws = self.active_workbook.active
        result = []
        if type(ws[sheet_range]) is tuple:
            if type(ws[sheet_range][0]) is not Cell:
                for row in ws[sheet_range]:
                    temp = []
                    for cell in row:
                        temp.append(cell.value)
                    result.append(temp)
            else:
                temp = []
                for cell in ws[sheet_range]:
                    temp.append(cell.value)
                result.append(temp)
            return result
        else:
            return ws[sheet_range].value

    def get_all_tables(self):
        ws = self.active_workbook.active
        return ws.tables.items()

    def get_table_data(self, table_name):
        ws = self.active_workbook.active
        table = ws.tables[table_name]
        return self.get_sheet_value(table.ref)

    def extend_table_body(self, table_name, number_of_rows=1):
        ws = self.active_workbook.active
        new_rows_ref = []
        while number_of_rows >= 1:
            try:
                table = ws.tables[table_name]
                tbl_ref = table.ref.split(":")
                old_col_ref = int(re.findall(r'\d+', tbl_ref[1])[0])
                row_ref = int(re.findall(r'\d+', tbl_ref[0])[0])
                ws.insert_rows(old_col_ref + 1, 1)
                new_col_ref = tbl_ref[1].replace(str(old_col_ref), str(old_col_ref + 1))
                row_ref = tbl_ref[0].replace(str(row_ref), str(old_col_ref + 1))
                tbl_ref[1] = new_col_ref
                new_tbl_ref = ":".join(tbl_ref)
                new_rows_ref.append(":".join([row_ref, new_col_ref]))
                table.ref = new_tbl_ref
                number_of_rows -= 1
            except KeyError:
                print("NAME ERROR: Table name not found")
                print("HINT: Tables present in sheet: " + str(self.get_all_tables()))
                break
        if len(new_rows_ref) > 1:
            return new_rows_ref
        elif len(new_rows_ref) == 1:
            return new_rows_ref[0]
        else:
            raise KeyError("Terminated. Table name invalid")

    def copy_data(self, sheet_range, save_alias,
                  datetime_fmt="%m/%d/%Y", true_value=True, entire_col=False):
        data = []
        ws = self.active_workbook.active
        boundaries = range_boundaries(sheet_range)
        min_col, min_row = boundaries[0], boundaries[1]
        max_col, max_row = boundaries[2], boundaries[3]

        if entire_col is True:
            max_row = None

        for row in ws.iter_rows(min_row=min_row, min_col=min_col,
                                max_row=max_row, max_col=max_col,
                                values_only=true_value):
            temp = []
            for cell in row:
                if type(cell) is datetime:
                    temp.append(cell.strftime(datetime_fmt))
                else:
                    temp.append(cell)
            data.append(temp)
        self.save_storage[save_alias] = data
        print(f"Data saved to storage. Alias: {save_alias}.")

    def paste_data(self, sheet_range, save_alias, overwrite=False, entire_col=False, dereference_when_pasted=True):
        ws = self.active_workbook.active
        data_set = self.save_storage[save_alias]
        if type(ws[sheet_range]) is Cell and entire_col is True:
            # Q2 (entire col) <- Goal
            col_letter = ws[sheet_range].column_letter
            maximum_row = ws.max_row
            sheet_range = ":".join([sheet_range, col_letter + str(maximum_row)])
        elif type(ws[sheet_range]) is tuple and entire_col is True:
            # Q2:R2 <- Goal
            maximum_row = ws.max_row
            temp = sheet_range.split(":")[1].replace(re.findall(r"\d+", sheet_range.split(":")[1])[0], str(maximum_row))
            sheet_range = ":".join([sheet_range.split(":")[0], temp])

        counter = 0
        for cells in ws[sheet_range]:
            if counter < len(data_set):
                for cell in cells:
                    self.insert_data_to_row(cell.coordinate, data_set[counter], overwrite=overwrite)
            counter += 1

        if dereference_when_pasted is True:
            del (self.save_storage[save_alias])

    def insert_data_to_row(self, row_ref, data,
                           overwrite=False):
        ws = self.active_workbook.active
        counter = 0
        if ":" in row_ref:
            for cell in ws[row_ref][0]:
                if '?fn?' not in str(data[counter]):
                    cell.value = data[counter]
                else:
                    # cel.value = "Expect a formulae to be here"
                    cell.value = self.__replicate_formulae(cell)
                counter += 1
        else:
            cell = ws[row_ref]
            if cell.value is None:
                if '?fn?' not in str(data[0]):
                    cell.value = data[0]
                else:
                    # cel.value = "Expect a formulae to be here"
                    cell.value = self.__replicate_formulae(cell)
            elif cell.value is not None and overwrite is True:
                if '?fn?' not in str(data[0]):
                    cell.value = data[0]
                else:
                    # cel.value = "Expect a formulae to be here"
                    cell.value = self.__replicate_formulae(cell)
            elif cell.value is not None and overwrite is False:
                pass

    def __replicate_formulae(self, cell_ref, reference_row_pos="ABOVE"):
        ws = self.active_workbook.active
        formulae_found = True

        if reference_row_pos == "ABOVE":
            row_num = cell_ref.row
            reference_row_num = row_num - 1
            reference_cell = cell_ref.coordinate.replace(str(row_num), str(reference_row_num))
            if "=" in ws[reference_cell].value:
                reference_cell_ref = cell_ref.coordinate.replace(str(row_num), str(row_num - 1))
            else:
                print(f"ERROR: The reference row ABOVE {cell_ref} does not contains formulae")
                formulae_found = False
        elif reference_row_pos == "BELOW":
            row_num = cell_ref.row
            reference_row_num = row_num + 1
            reference_cell = cell_ref.coordinate.replace(str(row_num), str(reference_row_num))
            if "=" in ws[reference_cell].value:
                reference_cell_ref = cell_ref.coordinate.replace(str(row_num), str(row_num + 1))
            else:
                print(f"ERROR: The reference row BELOW {cell_ref} does not contains formulae")
                formulae_found = False

        if formulae_found is True:
            # dissect formulae here
            replica_reference = ws[reference_cell_ref].value
            if "LOOKUP" in replica_reference.upper():
                return replica_reference.replace(str(reference_row_num), str(row_num), 1)
            else:
                return replica_reference.replace(str(reference_row_num), str(row_num))
        else:
            return "\'Error replicating formulae"

    def apply_formulae(self, formulae, sheet_range, entire_col=False, entire_row=False, has_header=False):
        pass

    def copy_formulae(self):
        pass

    def create_chart_line(self,
                          target_location,
                          chart_title=None,
                          chart_x_title=None,
                          chart_y_title=None,
                          reference_data_range=None,
                          chart_style=1):
        ws = self.active_workbook.active
        chart = LineChart()
        if type(ws[target_location]) is not Cell:
            print("The target location for the chart should be a single cell")
            return
        if reference_data_range is None:
            print("The range which will make up the chart data must be specified")
            return
        else:
            data_boundaries = range_boundaries(reference_data_range)
            data = Reference(ws, min_col=data_boundaries[0], min_row=data_boundaries[1],
                             max_col=data_boundaries[2], max_row=data_boundaries[3])
            chart.add_data(data, titles_from_data=True)
        if chart_title is not None:
            chart.title = chart_title
        if chart_x_title is not None:
            chart.x_axis.title = chart_x_title
        if chart_y_title is not None:
            chart.y_axis.title = chart_y_title

        chart.style = chart_style
        chart.marker = True
        chart.smooth = True
        chart_series = chart.series
        for line in chart_series:
            line.marker.symbol = "circle"

        ws.add_chart(chart, target_location)
        return chart

    def create_chart_bar(self,
                         target_location,
                         chart_title=None,
                         chart_x_title=None,
                         chart_y_title=None,
                         reference_data_range=None,
                         reference_category_range=None,
                         chart_type="col",
                         chart_style=1,
                         chart_shape=1,
                         chart_grouping=None,
                         chart_overlap=None):
        ws = self.active_workbook.active
        chart = BarChart()
        chart.type = chart_type
        if type(ws[target_location]) is not Cell:
            print("The target location for the chart should be a single cell")
            return
        if reference_data_range is None:
            print("The range which will make up the chart data must be specified")
            return
        else:
            data_boundaries = range_boundaries(reference_data_range)
            data = Reference(ws, min_col=data_boundaries[0], min_row=data_boundaries[1],
                             max_col=data_boundaries[2], max_row=data_boundaries[3])
            chart.add_data(data, titles_from_data=True)

        if reference_category_range is not None:
            category_boundary = range_boundaries(reference_category_range)
            cats = Reference(ws, min_col=category_boundary[0], min_row=category_boundary[1],
                             max_col=category_boundary[2], max_row=category_boundary[3])
            chart.set_categories(cats)
        if chart_title is not None:
            chart.title = chart_title
        if chart_x_title is not None:
            chart.x_axis.title = chart_x_title
        if chart_y_title is not None:
            chart.y_axis.title = chart_y_title
        if chart_overlap is not None:
            chart.overlap = chart_overlap
        if chart_grouping is not None:
            chart.grouping = chart_grouping

        chart.style = chart_style
        chart.shape = chart_shape

        ws.add_chart(chart, target_location)
        return chart

    def copy_chart(self, chart, save_alias):
        chart = deepcopy(chart)
        self.save_storage[save_alias] = chart
        print(f"Chart has been copied with alias: {save_alias}")

    def paste_chart(self, save_alias, target_location=None, auto_paste=True):
        ws = self.active_workbook.active
        if save_alias not in self.save_storage.keys():
            print("Save alias does not exist")
            return
        if type(ws[target_location]) is not Cell and auto_paste is True:
            print("The target location for the chart should be a single cell")
            return
        if auto_paste is False:
            return self.save_storage[save_alias]
        else:
            ws.add_chart(self.save_storage[save_alias], target_location)

    def delete_rows(self, row_number, amount=1):
        ws = self.active_workbook.active
        ws.delete_rows(row_number, amount=amount)

    def add_sort_and_filter(self, column):
        pass

    def remove_blanks_from_column(self, column_range,
                                  entire_column=False):
        ws = self.active_workbook.active
        empty_rows = []
        boundaries = list(range_boundaries(column_range))
        if entire_column is True:
            boundaries[3] = None
        col_data = ws.iter_rows(min_col=boundaries[0], min_row=boundaries[1],
                                max_col=boundaries[2], max_row=boundaries[3], values_only=False)
        for cell in col_data:
            data = cell[0]
            if data.value is None:
                empty_rows.append(data.row)

        for row in reversed(empty_rows):
            self.delete_rows(row)

    def remove_duplicates_from_column(self, column_range,
                                      entire_column=False, skip_empty_cell=True, save_alias=""):
        ws = self.active_workbook.active
        result = set()
        boundaries = list(range_boundaries(column_range))
        if entire_column is True:
            boundaries[3] = None
        col_data = ws.iter_rows(min_col=boundaries[0], min_row=boundaries[1],
                                max_col=boundaries[2], max_row=boundaries[3], values_only=True)
        for cell in col_data:
            data = cell[0]
            if data is None:
                if skip_empty_cell is not True:
                    result.add(data)
            else:
                result.add(data)

        if save_alias == "":
            return result
        else:
            self.save_storage[save_alias] = result

    def split_strings(self, data, delimiter, take_index=0, save_alias=""):
        result = []
        for strtxt in data:
            result.append(str(strtxt).split(delimiter)[take_index])

        return result
